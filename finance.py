from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
def Total_net_sales(need):
	while True:
		ws['B1'] = 2020
		ws['C1'] = 2019
		ws['D1'] = 2018
		ws['E1'] = 2017
		ws['F1'] = 2016
		TNS =input('是否輸入淨銷售額?:')
		if TNS == 'Y':
			TNS2020 = input('請輸入2020淨銷售額:') # Rows can also be appended
			TNS2019 = input('請輸入2019淨銷售額:')
			TNS2018 = input('請輸入2018淨銷售額:')
			TNS2017 = input('請輸入2017淨銷售額:')
			TNS2016 = input('請輸入2016淨銷售額:')
			ws.append(["Total net sales", TNS2020, TNS2019, TNS2018, TNS2017, TNS2016 ])
		else:
			break
def Net_income(need):
	while True:
		Ni =input('是否輸入淨利?:')
		if Ni == 'Y':
			Ni2020 = input('請輸入2020淨利:') # Rows can also be appended
			Ni2019 = input('請輸入2019淨利:')
			Ni2018 = input('請輸入2018淨利:')
			Ni2017 = input('請輸入2017淨利:')
			Ni2016 = input('請輸入2016淨利:')
			ws.append(["Net income", Ni2020, Ni2019, Ni2018, Ni2017, Ni2016 ])
		else :
			break	
def Earnings_per_share(need):
	while True:
		ws['A4'] = 'Earnings per share'
		EPS =input('是否輸入每股盈餘?:')
		if EPS == 'Y':
			EPSB =input('是否輸入基本每股盈餘?:')
			if EPSB == 'Y': 
				EPSB2020 = input('請輸入2020基本每股盈餘:') 
				EPSB2019 = input('請輸入2019基本每股盈餘:')
				EPSB2018 = input('請輸入2018基本每股盈餘:')
				EPSB2017 = input('請輸入2017基本每股盈餘:')
				EPSB2016 = input('請輸入2016基本每股盈餘:')
				ws.append(["Basic", EPSB2020, EPSB2019, EPSB2018, EPSB2017, EPSB2016 ])
			else :
				break
			EPSD =input('是否輸入2020稀釋每股盈餘?:')
			if EPSD == 'Y':
				EPSD2020 = input('請輸入2020稀釋每股盈餘:')
				EPSD2019 = input('請輸入2019稀釋每股盈餘:')
				EPSD2018 = input('請輸入2018稀釋每股盈餘:')
				EPSD2017 = input('請輸入2017稀釋每股盈餘:')
				EPSD2016 = input('請輸入2016稀釋每股盈餘:')
				ws.append(["Diluted", EPSD2020, EPSD2019, EPSD2018, EPSD2017, EPSD2016])
			else :
				break
		else :	
			break
def Cash_dividends_declared_per_share(need):
	while True:
		cashdivps =input('是否輸入每股股息?:')
		if cashdivps == 'Y':
			cdps2020 = input('請輸入2020每股股息')
			cdps2019 = input('請輸入2019每股股息')
			cdps2018 = input('請輸入2018每股股息')
			cdps2017 = input('請輸入2017每股股息')
			cdps2016 = input('請輸入2016每股股息')
			ws.append(["Cash dividends declared per share", cdps2020, cdps2019, cdps2018, cdps2017, cdps2016])
		else :
			break	
def Shares_used_in_computing_earnings_per_share(need):
	while True:
		ws['A8'] = 'Shares used in computing earnings per share'
		CEPS =input('是否輸入用於計算每股盈餘的股票?:')
		if CEPS == 'Y':
			CEPSB =input('是否輸入基本盈餘股數?:')
			if CEPSB == 'Y': 
				CEPSB2020 = input('請輸入2020基本盈餘股數:') 
				CEPSB2019 = input('請輸入2019基本盈餘股數:')
				CEPSB2018 = input('請輸入2018基本盈餘股數:')
				CEPSB2017 = input('請輸入2017基本盈餘股數:')
				CEPSB2016 = input('請輸入2016基本盈餘股數:')
				ws.append(["Basic", CEPSB2020, CEPSB2019, CEPSB2018, CEPSB2017, CEPSB2016 ])
			else :
				break
			CEPSD =input('是否輸入稀釋盈餘股數?:')
			if CEPSD == 'Y':
				CEPSD2020 = input('請輸入2020稀釋盈餘股數:')
				CEPSD2019 = input('請輸入2019稀釋盈餘股數:')
				CEPSD2018 = input('請輸入2018稀釋盈餘股數:')
				CEPSD2017 = input('請輸入2017稀釋盈餘股數:')
				CEPSD2016 = input('請輸入2016稀釋盈餘股數:')
				ws.append(["Diluted", CEPSD2020, CEPSD2019, CEPSD2018, CEPSD2017, CEPSD2016])
			else :
				break
		else :	
			break
def Totalcash_cashequivalents_marketablesecurities(need):
	while True:
		Totalcash =input('是否輸入現金`現金等價物及有價證券總額?:')
		if Totalcash == 'Y':
			Tc2020 = input('請輸入2020總額:')
			Tc2019 = input('請輸入2019總額:')
			Tc2018 = input('請輸入2018總額:')
			Tc2017 = input('請輸入2017總額:')
			Tc2016 = input('請輸入2016總額:')
			ws.append(["Total cash, cash equivalents and marketable securities", Tc2020, Tc2019, Tc2018, Tc2017, Tc2016])
		else :
			break
def Total_assets(need):
	while True:
		Ta =input('是否輸入總資產?:')
		if Ta == 'Y':
			Ta2020 = input('請輸入2020總資產:') # Rows can also be appended
			Ta2019 = input('請輸入2019總資產:')
			Ta2018 = input('請輸入2018總資產:')
			Ta2017 = input('請輸入2017總資產:')
			Ta2016 = input('請輸入2016總資產:')
			ws.append(["Total assets", Ta2020, Ta2019, Ta2018, Ta2017, Ta2016 ])
		else :
			break	
def Non_current_portion_of_term_debt(need):
	while True:
		NC =input('是否輸入非流動定期借款?:')
		if NC == 'Y':
			NC2020 = input('請輸入2020非流動定期借款:') # Rows can also be appended
			NC2019 = input('請輸入2019非流動定期借款:')
			NC2018 = input('請輸入2018非流動定期借款:')
			NC2017 = input('請輸入2017非流動定期借款:')
			NC2016 = input('請輸入2016非流動定期借款:')
			ws.append(["Non-current portion of term debt", NC2020, NC2019, NC2018, NC2017, NC2016 ])
		else :
			break
def Other_noncurrent_liabilities(need):
	while True:
		ONC =input('是否輸入其他非流動負債?:')
		if ONC == 'Y':
			ONC2020 = input('請輸入2020其他非流動負債:') # Rows can also be appended
			ONC2019 = input('請輸入2019其他非流動負債:')
			ONC2018 = input('請輸入2018其他非流動負債:')
			ONC2017 = input('請輸入2017其他非流動負債:')
			ONC2016 = input('請輸入2016其他非流動負債:')
			ws.append(["Other non-current liabilities", ONC2020, ONC2019, ONC2018, ONC2017, ONC2016 ])
		else :
			break
Total_net_sales(True)
Net_income(True)
Earnings_per_share(True)
Cash_dividends_declared_per_share(True)
Shares_used_in_computing_earnings_per_share(True)
Totalcash_cashequivalents_marketablesecurities(True)
Total_assets(True)
Non_current_portion_of_term_debt(True)
Other_noncurrent_liabilities(True)
# Save the file
wb.save("apple.xlsx")